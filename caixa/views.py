"""
Caixa - Views de conferência, conciliação e exportação analítica.

Responsabilidades:
- Tela de conferência por turno (comparativo PDV vs Stone vs iFood)
- Exportação de relatório analítico em Excel (5 abas)
- Drill-down de diferenças
"""

import io
import pandas as pd
from datetime import timedelta
from decimal import Decimal

from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import (
    Sum, Min, Max, Case, When, Value, F, Q, Subquery, OuterRef,
    CharField, DecimalField
)
from django.db.models.functions import Lower, Trim, Coalesce
from openpyxl.styles import PatternFill

from importacoes.models import VendaSWFast, TransacaoStone, PedidoIFood, FormaPagamento
from core.models import Empresa
from caixa.models import MovimentoCaixa


# ==========================================
# FUNÇÕES AUXILIARES
# ==========================================


def carregar_lojas(usuario):
    """Carrega apenas as lojas que o usuário logado tem permissão para ver."""
    try:
        if hasattr(usuario, 'perfil'):
            lojas_user = usuario.perfil.get_lojas()
            lojas_ids = lojas_user.values_list('id_empresa', flat=True)
            lojas_importadas = list(
                VendaSWFast.objects.filter(codigo_loja__isnull=False)
                .values_list('codigo_loja', flat=True)
                .distinct()
                .order_by('codigo_loja')
            )
            lojas_importadas = [str(l) for l in lojas_importadas]

            # Filtra apenas lojas que estão nas permitidas do usuário
            # Aqui buscamos a correspondência entre codigo_loja (SWFast) e id_empresa
            from core.models import Empresa
            codigos_permitidos = list(
                Empresa.objects.filter(id_empresa__in=lojas_ids)
                .values_list('ncad_swfast', flat=True)
            )
            codigos_permitidos = [str(c) for c in codigos_permitidos if c]
            return [loja for loja in lojas_importadas if loja in codigos_permitidos]

        return []
    except Exception as e:
        print(f"Erro ao carregar lojas: {e}")
        return []


def carregar_aberturas(codigo_loja):
    """Retorna lista de dicionários com número do caixa, data e turno."""
    if not codigo_loja:
        return []

    try:
        aberturas = (
            VendaSWFast.objects
            .filter(codigo_loja=codigo_loja)
            .exclude(Q(nr_abertura__isnull=True) | Q(nr_abertura=''))
            .values('nr_abertura')
            .annotate(min_data=Min('data_hora_transacao'))
            .order_by('-nr_abertura')
        )

        aberturas_formatadas = []
        for ab in aberturas:
            nr_abertura = str(ab['nr_abertura'])
            min_data = ab['min_data']
            texto_exibicao = f"Caixa: {nr_abertura}"

            if min_data:
                data_br = min_data.strftime('%d/%m/%Y')
                turno = "Dia" if 6 <= min_data.hour < 18 else "Noite"
                texto_exibicao = f"Caixa: {nr_abertura} - {data_br} - {turno}"

            aberturas_formatadas.append({
                'numero': nr_abertura,
                'rotulo': texto_exibicao
            })

        return aberturas_formatadas
    except Exception as e:
        print(f"Erro ao carregar aberturas: {e}")
        return []


def _get_periodo_stone(codigo_loja, nr_abertura):
    """Retorna (dt_inicio, dt_fim) com +1h de ajuste para consulta Stone."""
    agg = (
        VendaSWFast.objects
        .filter(codigo_loja=codigo_loja, nr_abertura=nr_abertura)
        .aggregate(
            min_abert=Min('dthr_abert_cx'),
            max_encerr=Max('dthr_encerr_cx')
        )
    )
    dt_inicio = agg['min_abert']
    dt_fim = agg['max_encerr']
    if dt_inicio:
        dt_inicio = dt_inicio + timedelta(hours=1)
    if dt_fim:
        dt_fim = dt_fim + timedelta(hours=1)
    return dt_inicio, dt_fim


def _get_empresa(codigo_loja):
    """Retorna a Empresa correspondente ao codigo_loja."""
    return Empresa.objects.filter(ncad_swfast=codigo_loja).first()


def _vendas_abertura(codigo_loja, nr_abertura):
    """QuerySet base das vendas de uma abertura."""
    return VendaSWFast.objects.filter(
        codigo_loja=codigo_loja,
        nr_abertura=nr_abertura
    )


def _ids_pedidos_ifood(codigo_loja, nr_abertura):
    """Lista de id_pedido_externo (lower/strip) do iFood nesta abertura."""
    return list(
        VendaSWFast.objects
        .filter(
            codigo_loja=codigo_loja,
            nr_abertura=nr_abertura,
            aplicativo__iexact='ifood'
        )
        .exclude(Q(id_pedido_externo__isnull=True) | Q(id_pedido_externo=''))
        .annotate(id_clean=Lower(Trim('id_pedido_externo')))
        .values_list('id_clean', flat=True)
    )


def buscar_dados_conferencia(codigo_loja, nr_abertura, incluir_dinheiro=False):
    """Busca dados de conciliação cruzando PDV, Stone e iFood."""
    try:
        empresa = _get_empresa(codigo_loja)
        vendas_qs = _vendas_abertura(codigo_loja, nr_abertura)

        if not incluir_dinheiro:
            vendas_qs = vendas_qs.exclude(forma_pagamento__iexact='DINHEIRO')

        # Agregar vendas SWFast por especific_form_pgto
        resumo_swfast = (
            vendas_qs
            .annotate(
                fp_lower=Lower(Trim('forma_pagamento')),
                app_lower=Lower(Trim('aplicativo'))
            )
            .values('fp_lower', 'app_lower', 'codigo_loja')
        )

        # Buscar forma de pagamento normalizada via Python (JOIN equivalente)
        # Agrupar vendas por especific_form_pgto
        from collections import defaultdict
        agrupado = defaultdict(Decimal)

        # Cache de FormaPagamento para esta loja
        fps = {}
        for fp in FormaPagamento.objects.filter(codigo_loja=codigo_loja):
            chave = (fp.forma_pagamento.strip().lower(), fp.codigo_loja, (fp.aplicativo or '').strip().lower())
            fps[chave] = fp.especific_form_pgto

        for venda in vendas_qs.values('forma_pagamento', 'aplicativo', 'codigo_loja', 'valor_pagamento'):
            chave = (
                venda['forma_pagamento'].strip().lower(),
                venda['codigo_loja'],
                (venda['aplicativo'] or '').strip().lower()
            )
            especific = fps.get(chave, 'OUTROS')
            agrupado[especific] += venda['valor_pagamento'] or Decimal('0')

        # Calcular Rel_importados para cada grupo
        resultados = []
        ids_ifood = _ids_pedidos_ifood(codigo_loja, nr_abertura)
        dt_inicio, dt_fim = _get_periodo_stone(codigo_loja, nr_abertura)

        for especific, valor_swfast in agrupado.items():
            rel_importados = Decimal('0')

            if especific == 'CARTAO':
                # Stone bruto - iFood total_pedido (pedidos integrados)
                valor_stone = Decimal('0')
                if empresa and dt_inicio and dt_fim:
                    agg_stone = (
                        TransacaoStone.objects
                        .filter(
                            stonecode=str(empresa.ncad_cartoes),
                            data_venda__range=(dt_inicio, dt_fim)
                        )
                        .aggregate(total=Sum('valor_bruto'))
                    )
                    valor_stone = agg_stone['total'] or Decimal('0')

                valor_ifood_cartao = Decimal('0')
                if empresa and ids_ifood:
                    pedidos_ifood = PedidoIFood.objects.filter(
                        id_pedido__in=[pid.upper() for pid in ids_ifood] if False else [],
                        id_restaurante=str(empresa.ncad_ifood)
                    )
                    # Usar lower/trim match
                    pedidos_ifood = PedidoIFood.objects.filter(
                        id_restaurante=str(empresa.ncad_ifood)
                    ).annotate(
                        id_clean=Lower(Trim('id_pedido'))
                    ).filter(id_clean__in=ids_ifood)
                    agg_ifood = pedidos_ifood.aggregate(total=Sum('total_pedido'))
                    valor_ifood_cartao = agg_ifood['total'] or Decimal('0')

                rel_importados = valor_stone - valor_ifood_cartao

            elif especific == 'PIX':
                rel_importados = Decimal('0')

            elif especific == 'IFOOD ONLINE':
                # Soma de vlr_pedido_sw dos pedidos iFood integrados que tem especific_form_pgto = IFOOD ONLINE
                if empresa and ids_ifood:
                    # Buscar pedidos iFood integrados nesta abertura com forma != dinheiro
                    # E cujas vendas SWFast associadas têm especific = IFOOD ONLINE
                    vendas_ifood_online_ids = []
                    for venda in _vendas_abertura(codigo_loja, nr_abertura).filter(aplicativo__iexact='ifood').exclude(
                        Q(id_pedido_externo__isnull=True) | Q(id_pedido_externo='')
                    ).values('id_pedido_externo', 'forma_pagamento', 'aplicativo', 'codigo_loja'):
                        chave = (
                            venda['forma_pagamento'].strip().lower(),
                            venda['codigo_loja'],
                            (venda['aplicativo'] or '').strip().lower()
                        )
                        if fps.get(chave) == 'IFOOD ONLINE':
                            vendas_ifood_online_ids.append(venda['id_pedido_externo'].strip().lower())

                    vendas_ifood_online_ids = list(set(vendas_ifood_online_ids))

                    if vendas_ifood_online_ids:
                        pedidos = (
                            PedidoIFood.objects
                            .filter(id_restaurante=str(empresa.ncad_ifood))
                            .annotate(id_clean=Lower(Trim('id_pedido')))
                            .filter(id_clean__in=vendas_ifood_online_ids)
                            .exclude(formas_pagamento__iexact='dinheiro')
                        )
                        agg = pedidos.aggregate(total=Sum('vlr_pedido_sw'))
                        rel_importados = agg['total'] or Decimal('0')

            elif especific == 'DINHEIRO':
                mov = MovimentoCaixa.objects.filter(
                    codigo_loja=codigo_loja,
                    nr_abertura=nr_abertura
                ).first()
                rel_importados = mov.valor_dinheiro_envelope if mov else Decimal('0')

            resultados.append({
                'RESUMO': especific,
                'rel_swfast': float(valor_swfast),
                'Rel_importados': float(rel_importados)
            })

        df = pd.DataFrame(resultados) if resultados else pd.DataFrame()

        # Pedidos cancelados
        df_cancelados = pd.DataFrame()
        if ids_ifood:
            cancelados = (
                PedidoIFood.objects
                .annotate(id_clean=Lower(Trim('id_pedido')))
                .filter(id_clean__in=ids_ifood)
                .exclude(Q(origem_cancelamento__isnull=True) | Q(origem_cancelamento=''))
                .values('nr_pedido', 'data', 'formas_pagamento', 'total_pedido')
            )
            if cancelados.exists():
                df_cancelados = pd.DataFrame(list(cancelados))

        # Incentivo iFood dinheiro
        valor_incentivo = Decimal('0')
        if ids_ifood:
            agg_inc = (
                PedidoIFood.objects
                .annotate(id_clean=Lower(Trim('id_pedido')))
                .filter(id_clean__in=ids_ifood, formas_pagamento__iexact='dinheiro')
                .aggregate(total=Sum('incentivo_ifood'))
            )
            valor_incentivo = float(agg_inc['total'] or 0)

    except Exception as e:
        print(f"Erro na conferência: {e}")
        df, df_cancelados, valor_incentivo = pd.DataFrame(), pd.DataFrame(), 0.0

    return df, df_cancelados, valor_incentivo


# ==========================================
# VIEW DE CONFERÊNCIA DE CAIXA
# ==========================================

@login_required(login_url='login')
def pagina_conferencia(request):
    context = {
        'lojas': carregar_lojas(request.user),
        'codigo_loja_selecionada': None,
        'nr_abertura_selecionada': None,
        'aberturas': [],
        'resultados': None
    }

    if request.method == 'GET' and 'codigo_loja' in request.GET:
        codigo_loja = request.GET.get('codigo_loja')
        nr_abertura = request.GET.get('nr_abertura')
        context['codigo_loja_selecionada'] = codigo_loja
        context['aberturas'] = carregar_aberturas(codigo_loja)

        if nr_abertura:
            context['nr_abertura_selecionada'] = nr_abertura

            primeira_venda = VendaSWFast.objects.filter(
                codigo_loja=codigo_loja,
                nr_abertura=nr_abertura,
                data_hora_transacao__isnull=False
            ).order_by('data_hora_transacao').first()

            if primeira_venda and primeira_venda.data_hora_transacao:
                context['data_caixa_selecionada'] = primeira_venda.data_hora_transacao.strftime('%d/%m/%Y')

            df_res, df_canc, val_inc = buscar_dados_conferencia(
                codigo_loja, nr_abertura,
                request.GET.get('incluir_dinheiro') == 'on'
            )

            if not df_res.empty:
                df_res = pd.concat([df_res, pd.DataFrame({
                    "RESUMO": ["INCENTIVO IFOOD A RECEBER DINHEIRO"],
                    "rel_swfast": [0.0], "Rel_importados": [val_inc]
                })], ignore_index=True)

                df_res["Rel_importados"] = pd.to_numeric(df_res["Rel_importados"], errors='coerce').fillna(0)
                df_res["rel_swfast"] = pd.to_numeric(df_res["rel_swfast"], errors='coerce').fillna(0)
                df_res["diferenca"] = df_res["Rel_importados"] - df_res["rel_swfast"]

                df_res = pd.concat([df_res, pd.DataFrame({
                    "RESUMO": ["SUBTOTAL"],
                    "Rel_importados": [df_res["Rel_importados"].sum()],
                    "rel_swfast": [df_res["rel_swfast"].sum()],
                    "diferenca": [df_res["diferenca"].sum()]
                })], ignore_index=True)

                context['resultados'] = df_res.fillna('').to_dict('records')
                context['cancelados'] = df_canc.to_dict('records')

    return render(request, 'caixa/conferencia.html', context)


# ==========================================
# EXPORTAÇÃO DO RELATÓRIO ANALÍTICO (EXCEL)
# ==========================================

@login_required(login_url='login')
def exportar_analitico_excel(request):
    codigo_loja = request.GET.get('codigo_loja')
    nr_abertura = request.GET.get('nr_abertura')
    incluir_dinheiro = request.GET.get('incluir_dinheiro') == 'on'

    if not codigo_loja or not nr_abertura:
        messages.error(request, "Parâmetros inválidos para o relatório analítico.")
        return redirect('conferencia_caixa')

    # Definir turno
    primeira_venda = (
        VendaSWFast.objects
        .filter(
            codigo_loja=codigo_loja,
            nr_abertura=nr_abertura,
            data_hora_transacao__isnull=False
        )
        .aggregate(min_data=Min('data_hora_transacao'))
    )

    turno = "Indefinido"
    min_data = primeira_venda['min_data']
    if min_data:
        turno = "Dia" if 6 <= min_data.hour < 18 else "Noite"

    # 1. RESUMO DO CAIXA
    df_res, df_canc, val_inc = buscar_dados_conferencia(codigo_loja, nr_abertura, incluir_dinheiro)

    if not df_res.empty:
        df_res = pd.concat([df_res, pd.DataFrame({
            "RESUMO": ["INCENTIVO IFOOD A RECEBER DINHEIRO"],
            "rel_swfast": [0.0], "Rel_importados": [val_inc]
        })], ignore_index=True)

        df_res["Rel_importados"] = pd.to_numeric(df_res["Rel_importados"], errors='coerce').fillna(0)
        df_res["rel_swfast"] = pd.to_numeric(df_res["rel_swfast"], errors='coerce').fillna(0)
        df_res["diferenca"] = df_res["Rel_importados"] - df_res["rel_swfast"]

        df_res = pd.concat([df_res, pd.DataFrame({
            "RESUMO": ["SUBTOTAL"],
            "Rel_importados": [df_res["Rel_importados"].sum()],
            "rel_swfast": [df_res["rel_swfast"].sum()],
            "diferenca": [df_res["diferenca"].sum()]
        })], ignore_index=True)

        df_res.rename(columns={
            'RESUMO': 'Resumo (Forma Pagto)',
            'Rel_importados': 'Rel. Importados (Stone/iFood)',
            'rel_swfast': 'Rel. SWFast (Caixa)',
            'diferenca': 'Diferença'
        }, inplace=True)
        df_res.insert(0, 'Caixa Nº', nr_abertura)
        df_res.insert(1, 'Turno', turno)
    else:
        df_res = pd.DataFrame(columns=[
            'Caixa Nº', 'Turno', 'Resumo (Forma Pagto)',
            'Rel. Importados (Stone/iFood)', 'Rel. SWFast (Caixa)', 'Diferença'
        ])

    empresa = _get_empresa(codigo_loja)

    # Cache de FormaPagamento para esta loja
    fps = {}
    for fp in FormaPagamento.objects.filter(codigo_loja=codigo_loja):
        chave = (fp.forma_pagamento.strip().lower(), fp.codigo_loja, (fp.aplicativo or '').strip().lower())
        fps[chave] = fp.especific_form_pgto

    # 2. ANALÍTICO IFOOD
    vendas_ifood = (
        _vendas_abertura(codigo_loja, nr_abertura)
        .filter(aplicativo__iexact='ifood')
        .order_by('data_hora_transacao')
        .values(
            'data_hora_transacao', 'venda', 'id_pedido_externo',
            'forma_pagamento', 'aplicativo', 'codigo_loja', 'valor_pagamento'
        )
    )

    # Agrupar por pedido iFood
    from collections import defaultdict
    pedidos_agrup = defaultdict(lambda: {
        'data_hora': None, 'venda': None, 'id_pedido': None,
        'formas_sw': set(), 'valor_total': Decimal('0'),
        'online': Decimal('0'), 'voucher': Decimal('0'), 'outros': Decimal('0')
    })

    for v in vendas_ifood:
        key = (v['data_hora_transacao'], v['venda'], v['id_pedido_externo'])
        pedidos_agrup[key]['data_hora'] = v['data_hora_transacao']
        pedidos_agrup[key]['venda'] = v['venda']
        pedidos_agrup[key]['id_pedido'] = v['id_pedido_externo']
        pedidos_agrup[key]['formas_sw'].add(v['forma_pagamento'])
        valor = v['valor_pagamento'] or Decimal('0')
        pedidos_agrup[key]['valor_total'] += valor

        chave_fp = (
            v['forma_pagamento'].strip().lower(),
            v['codigo_loja'],
            (v['aplicativo'] or '').strip().lower()
        )
        especific = fps.get(chave_fp, '')
        if especific == 'IFOOD ONLINE':
            pedidos_agrup[key]['online'] += valor
        elif especific == 'IFOOD VOUCHER':
            pedidos_agrup[key]['voucher'] += valor
        else:
            pedidos_agrup[key]['outros'] += valor

    rows_ifood = []
    for key, dados in pedidos_agrup.items():
        id_pedido_ext = dados['id_pedido']
        # Buscar pedido iFood correspondente
        pedido_ifood = None
        if id_pedido_ext:
            pedido_ifood = (
                PedidoIFood.objects
                .annotate(id_clean=Lower(Trim('id_pedido')))
                .filter(id_clean=(id_pedido_ext or '').strip().lower())
                .first()
            )

        rows_ifood.append({
            'Data SWFast': dados['data_hora'],
            'Nº Venda SWFast': dados['venda'],
            'ID Pedido iFood': id_pedido_ext,
            'Forma Pgto iFood (Portal)': pedido_ifood.formas_pagamento if pedido_ifood else '',
            'Forma Pgto SWFast': ', '.join(sorted(dados['formas_sw'])),
            'Valor SWFast (Total)': float(dados['valor_total']),
            'Apenas ONLINE': float(dados['online']),
            'Apenas VOUCHER': float(dados['voucher']),
            'Outros (Dinheiro/Cartão)': float(dados['outros']),
            'iFood Repasse': float(pedido_ifood.vlr_pedido_sw) if pedido_ifood else 0,
            'iFood Incentivo': float(pedido_ifood.incentivo_ifood) if pedido_ifood else 0,
            'Diferença': float((pedido_ifood.vlr_pedido_sw if pedido_ifood else Decimal('0')) - dados['valor_total']),
            'Status iFood': (pedido_ifood.status_pedido or 'CONCLUIDO') if pedido_ifood else 'CONCLUIDO'
        })

    df_ifood = pd.DataFrame(rows_ifood) if rows_ifood else pd.DataFrame(columns=[
        'Data SWFast', 'Nº Venda SWFast', 'ID Pedido iFood',
        'Forma Pgto iFood (Portal)', 'Forma Pgto SWFast',
        'Valor SWFast (Total)', 'Apenas ONLINE', 'Apenas VOUCHER',
        'Outros (Dinheiro/Cartão)', 'iFood Repasse', 'iFood Incentivo',
        'Diferença', 'Status iFood'
    ])

    # 3. CARTÕES NO SWFAST
    vendas_cartao = (
        _vendas_abertura(codigo_loja, nr_abertura)
        .order_by('data_hora_transacao')
        .values('data_hora_transacao', 'venda', 'forma_pagamento', 'valor_pagamento',
                'aplicativo', 'codigo_loja')
    )

    rows_cartoes = []
    for v in vendas_cartao:
        chave_fp = (
            v['forma_pagamento'].strip().lower(),
            v['codigo_loja'],
            (v['aplicativo'] or '').strip().lower()
        )
        if fps.get(chave_fp) == 'CARTAO':
            rows_cartoes.append({
                'Data SWFast': v['data_hora_transacao'],
                'Nº Venda': v['venda'],
                'Forma Pgto SWFast': v['forma_pagamento'],
                'Valor Registrado no Caixa': float(v['valor_pagamento'] or 0)
            })

    df_cartoes_sw = pd.DataFrame(rows_cartoes) if rows_cartoes else pd.DataFrame(columns=[
        'Data SWFast', 'Nº Venda', 'Forma Pgto SWFast', 'Valor Registrado no Caixa'
    ])

    # 4. TRANSAÇÕES REAIS STONE
    dt_inicio, dt_fim = _get_periodo_stone(codigo_loja, nr_abertura)

    rows_stone = []
    if empresa and dt_inicio and dt_fim:
        transacoes_stone = (
            TransacaoStone.objects
            .filter(
                stonecode=str(empresa.ncad_cartoes),
                data_venda__range=(dt_inicio, dt_fim)
            )
            .order_by('data_venda')
            .values('data_venda', 'bandeira', 'produto', 'qtd_parcelas',
                    'valor_bruto', 'desconto_mdr', 'valor_liquido')
        )
        for t in transacoes_stone:
            rows_stone.append({
                'Data Venda Stone': t['data_venda'],
                'Bandeira': t['bandeira'],
                'Produto': t['produto'],
                'Parcelas': t['qtd_parcelas'],
                'Valor Bruto': float(t['valor_bruto'] or 0),
                'Taxa MDR': float(t['desconto_mdr'] or 0),
                'Valor Líquido': float(t['valor_liquido'] or 0)
            })

    df_stone = pd.DataFrame(rows_stone) if rows_stone else pd.DataFrame(columns=[
        'Data Venda Stone', 'Bandeira', 'Produto', 'Parcelas',
        'Valor Bruto', 'Taxa MDR', 'Valor Líquido'
    ])

    # 5. IFOOD NÃO INTEGRADO
    # IDs de pedido já integrados no SWFast
    ids_integrados = list(
        VendaSWFast.objects
        .exclude(Q(id_pedido_externo__isnull=True) | Q(id_pedido_externo=''))
        .annotate(id_clean=Lower(Trim('id_pedido_externo')))
        .values_list('id_clean', flat=True)
    )

    # Período: entre MIN(data_hora_transacao) da abertura atual e MIN da próxima
    min_data_atual = (
        _vendas_abertura(codigo_loja, nr_abertura)
        .aggregate(min_data=Min('data_hora_transacao'))
    )['min_data']

    # Próxima abertura
    try:
        nr_abertura_int = int(nr_abertura)
        proxima_abertura = (
            VendaSWFast.objects
            .filter(codigo_loja=codigo_loja)
            .exclude(Q(nr_abertura__isnull=True) | Q(nr_abertura=''))
            .annotate(min_data=Min('data_hora_transacao'))
            .values('nr_abertura', 'min_data')
            .order_by('nr_abertura')
        )
        # Filtrar aberturas maiores que a atual
        min_data_proxima = None
        for ab in proxima_abertura:
            try:
                if int(ab['nr_abertura']) > nr_abertura_int and ab['min_data']:
                    min_data_proxima = ab['min_data']
                    break
            except (ValueError, TypeError):
                continue
    except (ValueError, TypeError):
        min_data_proxima = None

    rows_nao_integrado = []
    if min_data_atual:
        pedidos_nao_integrado = (
            PedidoIFood.objects
            .annotate(id_clean=Lower(Trim('id_pedido')))
            .exclude(id_clean__in=ids_integrados)
            .filter(data__gte=min_data_atual)
        )
        if min_data_proxima:
            pedidos_nao_integrado = pedidos_nao_integrado.filter(data__lt=min_data_proxima)

        for p in pedidos_nao_integrado.order_by('data').values(
            'data', 'id_pedido', 'vlr_pedido_sw', 'incentivo_ifood'
        ):
            vlr = float(p['vlr_pedido_sw'] or 0)
            inc = float(p['incentivo_ifood'] or 0)
            rows_nao_integrado.append({
                'Data no iFood': p['data'],
                'ID Pedido iFood': p['id_pedido'],
                'iFood Repasse': vlr,
                'iFood Incentivo': inc,
                'Valor Total Oculto do Caixa': vlr + inc
            })

    df_nao_integrado = pd.DataFrame(rows_nao_integrado) if rows_nao_integrado else pd.DataFrame(columns=[
        'Data no iFood', 'ID Pedido iFood', 'iFood Repasse',
        'iFood Incentivo', 'Valor Total Oculto do Caixa'
    ])

    # GERAÇÃO DO EXCEL
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_res.to_excel(writer, sheet_name='Resumo do Caixa', index=False)
        df_ifood.to_excel(writer, sheet_name='Analítico iFood (Integrados)', index=False)
        df_nao_integrado.to_excel(writer, sheet_name='iFood NÃO Integrado (Furos)', index=False)
        df_cartoes_sw.to_excel(writer, sheet_name='Cartões Passados no Caixa', index=False)
        df_stone.to_excel(writer, sheet_name='Transações Reais Stone', index=False)

        # Cor vermelha para cancelados
        worksheet_ifood = writer.sheets['Analítico iFood (Integrados)']
        fundo_vermelho = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        coluna_status = None
        for idx, col_name in enumerate(df_ifood.columns):
            if col_name == "Status iFood":
                coluna_status = idx + 1
                break

        if coluna_status:
            for row_idx in range(2, len(df_ifood) + 2):
                celula_status = worksheet_ifood.cell(row=row_idx, column=coluna_status).value
                if celula_status and str(celula_status).upper() != 'CONCLUIDO':
                    for col_idx in range(1, len(df_ifood.columns) + 1):
                        worksheet_ifood.cell(row=row_idx, column=col_idx).fill = fundo_vermelho

    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    nome_arquivo = f'Analitico_Loja{codigo_loja}_Cx{nr_abertura}_Turno_{turno}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response
